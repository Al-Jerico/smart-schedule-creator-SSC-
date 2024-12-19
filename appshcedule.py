from PyQt5.QtCore import Qt, QDate
from PyQt5.QtWidgets import (QApplication, QWidget, QGridLayout, QPushButton, 
                           QFileDialog, QLineEdit, QCheckBox, QCalendarWidget, QHBoxLayout,
                           QMessageBox, QComboBox, QLabel, QFrame, QScrollArea)
from PyQt5.QtGui import QColor, QPalette, QIcon
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from PyQt5.QtWidgets import QDateEdit
import sys
import numpy as np
import os
import json
from OTToolsTrainRefract import plot_schedule, Scheduler
from updatingSchedule import update_schedule

requests = {}

CONFIG_FILE = 'config.json'

def load_config():
    print(f"Looking for config file at: {os.path.abspath(CONFIG_FILE)}")
    if not os.path.exists(CONFIG_FILE):
        print(f"Config file not found. Creating default config.")
        default_config = {
            "receptionists": ["Dario", "Steve", "Goeffrey", "Nury", "Jacqueline", "Eddie"],
            "supervisors": ["Ana Rose", "Michelle", "Yousra"],
            "on_call_receptionists": ["Zynk", "Spark", "Nova"]
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(default_config, f, indent=4)
        print(f"Default config created at {os.path.abspath(CONFIG_FILE)}")
    
    with open(CONFIG_FILE, 'r') as f:
        config = json.load(f)
    #print("Loaded config:", json.dumps(config, indent=4))
    return config

class MyApp(QWidget):
    def __init__(self):
        super().__init__()

        self.config = load_config()
        #self.receptionists = self.config["receptionists"]
        #self.supervisors = self.config["supervisors"]
        #self.on_call_receptionists = self.config["on_call_receptionists"]

        self.excel_file = None
        self.excel_data = None
        self.requests = {'days_off': {}, 'shift_preferences': {}, 'mid_shifts': {}}
        self.arrivals = []
        self.departures = []
        self.start_date = None
        self.schedule_viz = None
        self.current_week = 0

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Schedule Generator')

        layout = QGridLayout()
        self.setLayout(layout)

        # Left column - Controls
        left_column = QGridLayout()
        
        button = QPushButton('Select Excel File')
        button.clicked.connect(self.selectExcelFile)
        left_column.addWidget(button, 0, 0)

        # Create a temporary scheduler to get worker names
        temp_scheduler = Scheduler(self.config,[], [], None)
        all_workers = temp_scheduler.worker_names + temp_scheduler.supervisor_names

        # Worker selection combobox
        self.name_input = QComboBox()
        self.name_input.addItems(all_workers)
        left_column.addWidget(self.name_input, 1, 0)

        # Request type selection (Day Off or Shift Preference)
        self.request_type = QComboBox()
        self.request_type.addItems(['Day Off', 'Shift Preference', 'MID Shift'])
        left_column.addWidget(self.request_type, 2, 0)

        # Shift selection (AM/PM)
        self.shift_input = QComboBox()
        # Dans la méthode initUI de la classe MyApp
        self.shift_input.addItems(['AM', 'PM'])
        self.shift_input.setEnabled(False)  
        left_column.addWidget(self.shift_input, 3, 0)

        # Enable/disable shift selection based on request type
        self.request_type.currentTextChanged.connect(self.onRequestTypeChanged)

        # Start date selection
        start_date_label = QLabel('Start Date:')
        left_column.addWidget(start_date_label, 4, 0)
        
        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate())
        self.start_date_input.dateChanged.connect(self.on_start_date_changed)
        left_column.addWidget(self.start_date_input, 5, 0)

        self.calendar = QCalendarWidget()
        self.calendar.selectionChanged.connect(self.update_dates)
        left_column.addWidget(self.calendar, 6, 0)

        button = QPushButton('Add Request')
        button.clicked.connect(self.addRequest)
        left_column.addWidget(button, 7, 0)

        button = QPushButton('Generate Schedule')
        button.clicked.connect(self.generateSchedule)
        left_column.addWidget(button, 8, 0)

        button = QPushButton('Update Schedule')
        button.clicked.connect(self.updateScheduleWindow)
        left_column.addWidget(button, 9, 0)

        # Create a horizontal layout for the arrow buttons
        self.week_nav_layout = QHBoxLayout()

        # Create left arrow button
        self.prev_week_button = QPushButton()
        self.prev_week_button.setIcon(QIcon.fromTheme("go-previous"))
        self.prev_week_button.clicked.connect(self.switch_week_backward)

        # Create right arrow button
        self.next_week_button = QPushButton()
        self.next_week_button.setIcon(QIcon.fromTheme("go-next"))
        self.next_week_button.clicked.connect(self.switch_week_forward)

        # Create a label to display current week
        self.week_label = QLabel("Week 1")
        self.week_label.setAlignment(Qt.AlignCenter)

        # Add widgets to the horizontal layout
        self.week_nav_layout.addWidget(self.prev_week_button)
        self.week_nav_layout.addWidget(self.week_label)
        self.week_nav_layout.addWidget(self.next_week_button)

        # Create a widget to hold the horizontal layout
        week_nav_widget = QWidget()
        week_nav_widget.setLayout(self.week_nav_layout)

        # Add the week navigation widget to the main layout
        self.layout().addWidget(week_nav_widget, 1, 1)  # Add below the schedule visualization

        # Keep track of the current week
        self.current_week = 0

        # Add left column to main layout
        layout.addLayout(left_column, 0, 0)

        self.setAcceptDrops(True)
        self.resize(1200, 600)
        self.show()

    def switch_week_forward(self):
        total_weeks = (len(self.excel_data.columns) - 1) // 7
        print(total_weeks)
        self.current_week = (self.current_week + 1) % total_weeks
        self.update_week_label()
        self.init_schedule_viz()

    def switch_week_backward(self):
        total_weeks = (len(self.excel_data.columns) - 1) // 7
        self.current_week = (self.current_week - 1) % total_weeks
        self.update_week_label()
        self.init_schedule_viz()

    def addMidShift(self):
        worker = self.name_input.currentText()
        date = self.calendar.selectedDate().toPyDate()
        if worker not in self.requests['mid_shifts']:
            self.requests['mid_shifts'][worker] = []
        self.requests['mid_shifts'][worker].append(date)
        QMessageBox.information(self, "Success", f"MID shift added for {worker} on {date}")

    def update_week_label(self):
        self.week_label.setText(f"Week {self.current_week + 1}")

    def onRequestTypeChanged(self, text):
        # Enable shift selection only for shift preferences
        self.shift_input.setEnabled(text == 'Shift Preference')
        # Disable shift selection for MID Shift (it's always MID)
        if text == 'MID Shift':
            self.shift_input.setCurrentText('MID')
            self.shift_input.setEnabled(False)

    def on_start_date_changed(self, qdate):
        self.start_date = qdate.toPyDate() 
        if self.excel_data is not None:
            self.init_schedule_viz()

    def selectExcelFile(self):
        filePath, _ = QFileDialog.getOpenFileName(self, 'Select Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if filePath:
            self.excel_data = pd.read_excel(filePath)
            
            arrivals_index = None
            departures_index = None
            
            for index, row in self.excel_data.iterrows():
                if 'Arrivals' in row.values:
                    arrivals_index = index
                if 'Departures' in row.values:
                    departures_index = index
            
            if arrivals_index is not None and departures_index is not None:
                self.arrivals = self.excel_data.loc[arrivals_index].tolist()[1:]
                self.departures = self.excel_data.loc[departures_index].tolist()[1:]
                print(self.arrivals)
                print(self.departures)
            else:
                print("Error: Could not find 'Arrivals' or 'Departures' in the Excel file.")

            # If start date is set, initialize the schedule visualization
            if self.start_date:
                self.init_schedule_viz()

    def update_dates(self):
        # This is now only for the calendar widget selection
        selected_date = self.calendar.selectedDate()
        print(f"Selected date for request: {selected_date.toPyDate()}")

    def init_schedule_viz(self):
        # Remove existing visualization if it exists
        if self.schedule_viz:
            self.layout().removeWidget(self.schedule_viz)
            self.schedule_viz.deleteLater()
        
        # Calculate the start date for the current week
        current_start_date = self.start_date + timedelta(days=self.current_week * 7)
        print(current_start_date)
        
        # Create new visualization with start_date, requests, and current week
        self.schedule_viz = ScheduleVisualization(current_start_date, self.excel_data, self.requests, self.current_week)
        scroll = QScrollArea()
        scroll.setWidget(self.schedule_viz)
        scroll.setWidgetResizable(True)
        self.layout().addWidget(scroll, 0, 1)  # Add to right side (column 1)

        # Update the week label
        self.update_week_label()

    def update_week_label(self):
        self.week_label.setText(f"Semaine {self.current_week + 1}")

    def addRequest(self):
        worker = self.name_input.currentText()
        request_type = self.request_type.currentText()
        selected_date = self.calendar.selectedDate().toPyDate()

        if request_type == 'Day Off':
             # Remove any existing request for this date
            if worker in self.requests['days_off']:
                self.requests['days_off'][worker] = [
                    date for date in self.requests['days_off'][worker]
                    if date != selected_date
                ]
            if worker in self.requests['shift_preferences']:
                self.requests['shift_preferences'][worker] = [
                    (date, shift) for date, shift in self.requests['shift_preferences'][worker]
                    if date != selected_date
                ]

            self.requests['days_off'].setdefault(worker, []).append(selected_date)
            QMessageBox.information(self, 'Success', f'Added day off request for {worker} on {selected_date}')
        
        elif request_type == 'Shift Preference':
            shift = self.shift_input.currentText()

            # Remove any existing request for this date
            if worker in self.requests['days_off']:
                self.requests['days_off'][worker] = [
                    date for date in self.requests['days_off'][worker]
                    if date != selected_date
                ]
            if worker in self.requests['shift_preferences']:
                 self.requests['shift_preferences'][worker] = [
                    (date, shift) for date, shift in self.requests['shift_preferences'][worker]
                    if date != selected_date
                ]
            
            self.requests['shift_preferences'].setdefault(worker, []).append((selected_date, shift))
            QMessageBox.information(self, 'Success', f'Added {shift} shift preference for {worker} on {selected_date}')

        elif request_type == 'MID Shift':
            # Vérifier si le travailleur est un superviseur
            supervisors = ['Ana Rose', 'Michelle', 'Yousra']
            if worker in supervisors:
                QMessageBox.warning(self, "Error", f"MID shifts can only be assigned to receptionists, not to supervisors like {worker}.")
                return

            # Ajouter le MID shift
            if worker not in self.requests['mid_shifts']:
                self.requests['mid_shifts'][worker] = []
            self.requests['mid_shifts'][worker].append(selected_date)
            QMessageBox.information(self, 'Success', f'Added MID shift for {worker} on {selected_date}')

        print("Current requests:", self.requests)
        self.init_schedule_viz()  # Refresh the visualization


    @staticmethod
    def apply_mid_shifts(schedule_df, mid_shifts):
        for worker, dates in mid_shifts.items():
            for date in dates:
                date_str = date.strftime('%Y-%m-%d')
                if date_str in schedule_df.columns:
                    # Trouver un travailleur AM et PM à remplacer
                    am_workers = schedule_df[(schedule_df[date_str] == '(AM)') & 
                                         (~schedule_df['Worker'].isin(['Ana Rose', 'Michelle', 'Yousra', 'Boris']))]
                    pm_workers = schedule_df[(schedule_df[date_str] == '(PM)') & 
                                            (~schedule_df['Worker'].isin(['Ana Rose', 'Michelle', 'Yousra', 'Boris']))]
                    
                    if am_workers.empty or pm_workers.empty:
                        print(f"Impossible d'appliquer le MID shift pour {worker} le {date_str}. Pas assez de réceptionnistes disponibles.")
                        continue
                
                    am_worker = am_workers.iloc[0]['Worker']
                    pm_worker = pm_workers.iloc[0]['Worker']

                    # Compter les shifts pour la semaine
                    week_start = date - timedelta(days=date.weekday())
                    week_end = week_start + timedelta(days=6)
                    week_dates = [d.strftime('%Y-%m-%d') for d in pd.date_range(week_start, week_end)]
                    
                    am_shifts = schedule_df[schedule_df['Worker'] == am_worker][week_dates].eq('(AM)').sum().sum()
                    pm_shifts = schedule_df[schedule_df['Worker'] == pm_worker][week_dates].eq('(PM)').sum().sum()
                    
                    # Remplacer les shifts
                    schedule_df.loc[schedule_df['Worker'] == am_worker, date_str] = np.nan
                    schedule_df.loc[schedule_df['Worker'] == pm_worker, date_str] = np.nan
                    schedule_df.loc[schedule_df['Worker'] == worker, date_str] = '(MID)'
        
        return schedule_df


    def generateSchedule(self):
        try:
            # Get the directory where the executable is located
            if getattr(sys, 'frozen', False):
                # If running as executable
                exe_dir = os.path.dirname(sys.executable)
            else:
                # If running as script
                exe_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Create schedules directory if it doesn't exist
            schedules_dir = os.path.join(exe_dir, 'schedules')
            os.makedirs(schedules_dir, exist_ok=True)
            
            # Change to schedules directory
            os.chdir(schedules_dir)
            
            print(f"Current working directory: {os.getcwd()}")
            print(f"Start date: {self.start_date}")
            print(f"Arrivals length: {len(self.arrivals)}")
            print(f"Departures length: {len(self.departures)}")
            
            # Pass both days off and shift preferences to the scheduler
            plot_schedule(
                self.config,
                self.departures, 
                self.arrivals, 
                self.start_date, 
                self.requests['days_off'],
                shift_preferences=self.requests['shift_preferences'],
                prev_week_shifts=None,
                mid_shifts=self.requests['mid_shifts']
            )
            
            QMessageBox.information(self, "Success", f"Schedule generated successfully!\nFiles saved in: {schedules_dir}")


        except Exception as e:
            import traceback
            error_msg = f"Error generating schedule:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "Error", error_msg)

    def updateScheduleWindow(self):
        self.update_window = UpdateScheduleWindow()
        self.update_window.show()

class ScheduleVisualization(QFrame):  
    def __init__(self, start_date=None, excel_data=None, requests=None, current_week=0):
        super().__init__()
        self.setFrameStyle(QFrame.Box | QFrame.Plain)
        self.grid = QGridLayout()
        self.setLayout(self.grid)
        self.start_date = start_date
        self.current_week = current_week
        self.excel_data = excel_data
        self.requests = requests
        self.config = load_config()
        
        # Initialize default workers
        temp_scheduler = Scheduler(self.config, [], [], None)
        self.workers = temp_scheduler.worker_names + temp_scheduler.supervisor_names
        self.requests_by_day = {}
        
        # Update workers list if excel data is provided
        if excel_data is not None:
            excel_workers = []
            for index, row in excel_data.iterrows():
                if 'Arrivals' in row.values or 'Departures' in row.values:
                    continue
                if not pd.isna(row.iloc[0]):
                    excel_workers.append(row.iloc[0])
            if excel_workers:  # Only update if we found workers in Excel
                self.workers = excel_workers
        
        # Update requests
        if requests is not None:
            self.requests = requests
        else:
            self.requests = {'days_off': {}, 'shift_preferences': {}}
        
        self.init_grid()

    def init_grid(self):
        # Clear existing grid
        for i in reversed(range(self.grid.count())): 
            self.grid.itemAt(i).widget().setParent(None)
            
        if not self.start_date:
            return
            
        # Add headers
        self.grid.addWidget(QLabel('Worker'), 0, 0)
        
        # Calculate dates for the next 7 days
        dates = [self.start_date + timedelta(days=i) for i in range(7)]
        
        # Add date headers
        for col, date in enumerate(dates, start=1):
            date_label = QLabel(date.strftime('%Y-%m-%d'))
            self.grid.addWidget(date_label, 0, col)
            
        """
        
        # Add worker rows
        for row, worker in enumerate(self.workers, start=1):
            worker_label = QLabel(worker)
            self.grid.addWidget(worker_label, row, 0)
            
            # Add cells for each date
            for col, date in enumerate(dates, start=1):
                cell = QLabel()
                cell.setStyleSheet('border: 1px solid black; padding: 5px;')
                
                # Afficher les requêtes pour ce travailleur à cette date
                request_text = self.get_requests_for_date(worker, date)
                if request_text:
                    cell.setText(request_text)
                    cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #FFE4E1;')
                
                self.grid.addWidget(cell, row, col)

        """
        
        
        for i, worker in enumerate(self.workers, start=1):
            # ... (code existant pour afficher le nom du travailleur)
            worker_label = QLabel(worker)
            self.grid.addWidget(worker_label, i, 0)

            for j in range(7):
                date = self.start_date + timedelta(days=j)
                cell = QLabel()
                cell.setAlignment(Qt.AlignCenter)
                cell.setStyleSheet('border: 1px solid black; padding: 5px;')
                
                # Vérifier les requêtes pour cette date et ce travailleur
                day_off, shift_pref, mid_shift = self.get_requests_for_date(worker, date)
                
                if day_off:
                    cell.setText('OFF')
                    cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #FFB3BA;')
                elif shift_pref:
                    cell.setText(f'Préf: {shift_pref}')
                    cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #BAFFC9;')
                elif mid_shift:
                    cell.setText('MID')
                    cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #BAE1FF;')
                
                self.grid.addWidget(cell, i, j + 1)


    """
    
    def get_requests_for_date(self, worker, date):
        # Adjust the date based on the current week
        adjusted_date = date + timedelta(days=self.current_week * 7)
        
        requests = []
        
        # Check for days off
        if worker in self.requests['days_off']:
            if date in self.requests['days_off'][worker]:
                requests.append("Congé")
        
        # Check for shift preferences
        if worker in self.requests['shift_preferences']:
            for req_date, shift in self.requests['shift_preferences'][worker]:
                if req_date == date:
                    requests.append(f"Préf: {shift}")

        # Check for mid shifts
        if worker in self.requests['mid_shifts']:
            if adjusted_date in self.requests['mid_shifts'][worker]:
                requests.append("MID")
        
        return "\n".join(requests) if requests else ""


    """

    def get_requests_for_date(self, worker, date):
        day_off = False
        shift_pref = None
        mid_shift = False

        if worker in self.requests['days_off'] and date in self.requests['days_off'][worker]:
            day_off = True
        
        if worker in self.requests['shift_preferences']:
            for req_date, shift in self.requests['shift_preferences'][worker]:
                if req_date == date:
                    shift_pref = shift
                    break
        
        if worker in self.requests['mid_shifts'] and date in self.requests['mid_shifts'][worker]:
            mid_shift = True

        return day_off, shift_pref, mid_shift

    def update_cell(self, worker, date, shift):
        # Trouver la cellule correspondante et mettre à jour son contenu
        worker_row = self.workers.index(worker) + 1
        date_col = (date - self.start_date).days + 1
        
        if 0 <= worker_row < self.grid.rowCount() and 0 <= date_col < self.grid.columnCount():
            cell = self.grid.itemAtPosition(worker_row, date_col).widget()
            current_text = cell.text()
            
            if shift == 'MID':
                cell.setText('MID')
                cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #BAE1FF;')
            else:
                current_text = cell.text()
                
                if current_text:
                    cell.setText(f"{current_text}\nPréf: {shift}")
                else:
                    cell.setText(f"Préf: {shift}")
                
                cell.setStyleSheet('border: 1px solid black; padding: 5px; background-color: #BAFFC9;')

class UpdateScheduleWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Update Schedule')

        layout = QGridLayout()
        self.setLayout(layout)

        button = QPushButton('Select Existing Schedule')
        button.clicked.connect(self.selectExistingSchedule)
        layout.addWidget(button, 0, 0)

        button = QPushButton('Select New Arrivals and Departures')
        button.clicked.connect(self.selectNewArrivalsDepartures)
        layout.addWidget(button, 1, 0)

        button = QPushButton('Update Schedule')
        button.clicked.connect(self.updateSchedule)
        layout.addWidget(button, 2, 0)

        self.existing_schedule = None
        self.new_arrivals_departures = None

    def selectExistingSchedule(self):
        filePath, _ = QFileDialog.getOpenFileName(self, 'Select Existing Schedule', '', 'Excel Files (*.xlsx *.xls)')
        if filePath:
            self.existing_schedule = pd.read_excel(filePath)

    def selectNewArrivalsDepartures(self):
        filePath, _ = QFileDialog.getOpenFileName(self, 'Select New Arrivals and Departures', '', 'Excel Files (*.xlsx *.xls)')
        if filePath:
            self.new_arrivals_departures = pd.read_excel(filePath)

    def updateSchedule(self):
        if self.existing_schedule is not None and self.new_arrivals_departures is not None:
            styled_schedule, raw_schedule = update_schedule(self.existing_schedule, self.new_arrivals_departures)
            
            # Save the styled schedule
            with pd.ExcelWriter('updated_schedule.xlsx', engine='openpyxl') as writer:
                raw_schedule.to_excel(writer, sheet_name='Schedule', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Schedule']
                
                # Apply colors based on conditions
                for row_idx in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
                    for col_idx in range(2, worksheet.max_column + 1):  # Start from 2 to skip first column
                        current_cell = worksheet.cell(row=row_idx, column=col_idx)
                        prev_cell = worksheet.cell(row=row_idx, column=col_idx-1)
                        
                        # Check for short night (PM followed by AM)
                        if ('(PM)' in str(prev_cell.value) and 
                            '(AM)' in str(current_cell.value)):
                            current_cell.fill = openpyxl.styles.PatternFill(
                                start_color='FFCCCC',
                                end_color='FFCCCC',
                                fill_type='solid'
                            )
            
            QMessageBox.information(self, "Success", "Schedule updated successfully")
        else:
            QMessageBox.warning(self, "Error", "Please select both existing schedule and new arrivals and departures")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())